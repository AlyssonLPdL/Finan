from datetime import datetime
from flask import Flask, render_template, request, redirect,  url_for, flash, session, jsonify, send_file
import sqlite3
import io
import openpyxl
from openpyxl import Workbook
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import os
from io import BytesIO

app = Flask(__name__)
app.secret_key = 's3cr3tK3y'
UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Configuração inicial do banco de dados
def init_db():
    with sqlite3.connect('tmp/finance.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                quantia NUMERIC NOT NULL,
                description TEXT NOT NULL,
                value REAL NOT NULL,
                payment_method TEXT NOT NULL,
                type TEXT NOT NULL,
                user_id INTEGER,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL
            )
        ''')
        conn.commit()

init_db()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/', methods=['GET', 'POST'])
@login_required
def index():
    if request.method == 'POST':
        date = request.form['date']
        quantia = request.form['quantia']
        description = request.form['description']
        value = float(request.form['value'])
        payment_method = request.form['payment_method']
        type_ = request.form['type']

        user_id = session.get('user_id')
        if not user_id:
            flash('Você precisa estar logado para adicionar uma transação.', 'warning')
            return redirect(url_for('login'))

        if not date:
            date = datetime.today().strftime('%Y-%m-%d')  # Formato YYYY-MM-DD

        parcelas = request.form.get('parcelas')  # Pega o valor de parcelas
        if parcelas:
            parcelas = int(parcelas)
        else:
            parcelas = 1  # Valor padrão para parcelas se não for fornecido

        # Se o método de pagamento for Crédito, adicionar a informação das parcelas na descrição
        if payment_method in ['Credito_nu', 'Credito_c6']:
            description = f"{description} - x{parcelas}"

        # Se a quantia não for fornecida, usa o valor padrão de 1
        if not quantia:
            quantia = 1  # Valor padrão para quantia

        with sqlite3.connect('tmp/finance.db') as conn:
            cursor = conn.cursor()
            cursor.execute('''
            INSERT INTO transactions (date, quantia, description, value, payment_method, type, user_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (date, quantia, description, value, payment_method, type_, user_id))
            conn.commit()

        return redirect('/')

    search_query = request.args.get('search', '')  # Captura a busca inteligente
    month_year_filter = request.args.get('month_year')  # Captura o filtro de mês (se houver)
    payment_filter = request.args.get('payment_method', '')  # Pega o filtro da URL
    tipe_filter = request.args.get('type', '')
    payment_methods = ["Pix", "Dinheiro", "Credito_c6", "Credito_nu"]
    tipos = ["Gasto", "Ganho"]

    query = "SELECT * FROM transactions WHERE user_id = ?"
    params = [session.get('user_id')]

    if payment_filter:
        query += " AND payment_method = ?"
        params.append(payment_filter)
        
    if month_year_filter:
        query += " AND strftime('%Y-%m', date) = ?"
        params.append(month_year_filter)

    if tipe_filter:
        query += " AND type = ?"
        params.append(tipe_filter)

    # Implementação da busca inteligente
    if search_query:
        if search_query.startswith('>'):
            try:
                value = float(search_query[1:])  # Captura o valor após o `>`
                query += ' AND value > ?'
                params.append(value)
            except ValueError:
                pass  # Ignorar se não for um número válido
        elif search_query.startswith('<'):
            try:
                value = float(search_query[1:])  # Captura o valor após o `<`
                query += ' AND value < ?'
                params.append(value)
            except ValueError:
                pass  # Ignorar se não for um número válido
        else:
            # Busca em outros campos: descrição, data, método de pagamento
            query += '''
                AND (description LIKE ? OR 
                     date LIKE ? OR 
                     payment_method LIKE ? OR 
                     CAST(value AS TEXT) LIKE ?)
            '''
            search_term = f'%{search_query}%'  # Termo genérico
            params.extend([search_term, search_term, search_term, search_term])

    query += ' ORDER BY date'

    with sqlite3.connect('tmp/finance.db') as conn:
        cursor = conn.cursor()
        cursor.execute(query, params)
        transactions = cursor.fetchall()

        print(query)
        print(params)
        user_id = session.get('user_id')
        print(f"User ID: {user_id}")  # Verifique o valor de user_id

        # Obter todos os meses únicos disponíveis
        cursor.execute("SELECT DISTINCT strftime('%Y-%m', date) as month_year FROM transactions ORDER BY month_year")
        months_years = [row[0] for row in cursor.fetchall()]

    return render_template('index.html', transactions=transactions, months_years=months_years, month_year_filter=month_year_filter, payment_filter=payment_filter, payment_methods=payment_methods, tipos=tipos, tipe_filter=tipe_filter)

# Tela de login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        with sqlite3.connect('tmp/finance.db') as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT id, password FROM users WHERE email = ?', (email,))
            user = cursor.fetchone()

            if user and check_password_hash(user[1], password):
                session['user_id'] = user[0]
                flash('Login realizado com sucesso!', 'success')
                return redirect('/')
            else:
                flash('Email ou senha inválidos.', 'danger')

    return render_template('login.html')

# Tela de cadastro
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form['confirm_password']

        if password != confirm_password:
            flash('As senhas não coincidem.', 'danger')
            return redirect(url_for('register'))

        hashed_password = generate_password_hash(password)

        with sqlite3.connect('tmp/finance.db') as conn:
            cursor = conn.cursor()
            try:
                cursor.execute('INSERT INTO users (email, password) VALUES (?, ?)', (email, hashed_password))
                conn.commit()
                flash('Cadastro realizado com sucesso!', 'success')
                return redirect(url_for('login'))
            except sqlite3.IntegrityError:
                flash('O email já está em uso.', 'danger')

    return render_template('register.html')

# Logout
@app.route('/logout')
def logout():
    session.pop('user_id', None)
    flash('Você saiu da sua conta.', 'info')
    return redirect(url_for('login'))

@app.route('/generate_excel', methods=['GET'])
def generate_excel():
    # Conectar ao banco de dados e obter as transações
    with sqlite3.connect('tmp/finance.db') as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, date, quantia, description, value, payment_method, type FROM transactions")
        transactions = cursor.fetchall()

    # Criar um arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Transações"

    # Cabeçalho da planilha
    ws.append(["ID", "Data", "Quantia", "Descrição", "Valor", "Método de Pagamento", "Tipo"])

    # Adicionar os dados das transações
    for transaction in transactions:
        ws.append(transaction)

    # Salvar a planilha em memória (em vez de salvar fisicamente)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar a planilha para o cliente
    return send_file(output, as_attachment=True, download_name="transacoes.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def save_excel_to_db(file_path):
    try:
        print(f"Tentando abrir o arquivo: {file_path}")
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        print("Planilha carregada com sucesso.")
    except Exception as e:
        print(f"Erro ao carregar a planilha: {e}")
        raise

    # Obtenha o user_id da sessão
    user_id = session.get('user_id')
    if not user_id:
        print("Usuário não autenticado. Não é possível salvar as transações.")
        return

    # Verifique as primeiras linhas para garantir que os dados estão sendo lidos
    for row in ws.iter_rows(min_row=2, values_only=True):  # Pular o cabeçalho
        # Validar se a linha tem valores válidos
        if None in row or len(row) < 6:
            continue

        # Verificar se a transação já foi inserida (para evitar duplicações)
        date = row[0]
        description = row[2]
        cursor = sqlite3.connect('tmp/finance.db').cursor()
        cursor.execute('SELECT * FROM transactions WHERE date = ? AND description = ?', (date, description))
        existing_transaction = cursor.fetchone()

        if existing_transaction:
            continue  # Pular a inserção da transação duplicada

        try:
            with sqlite3.connect('tmp/finance.db') as conn:
                cursor = conn.cursor()
                cursor.execute(''' 
                    INSERT INTO transactions (date, quantia, description, value, payment_method, type, user_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', row[:6] + (user_id,))  # Inclui o user_id ao salvar
                conn.commit()
        except Exception as e:
            print(f"Erro ao inserir: {row} - {e}")
            continue  # Ignorar erro e continuar com a próxima linha

# Rota para upload da planilha
@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    file = request.files['file']
    if file and file.filename.endswith('.xlsx'):
        # Caminho do arquivo
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)

        # Verificar se o arquivo já existe e renomeá-lo se necessário
        if os.path.exists(file_path):
            base_name, ext = os.path.splitext(file.filename)
            counter = 1
            while os.path.exists(file_path):
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{base_name} ({counter}){ext}")
                counter += 1

        try:
            # Salvar o arquivo no diretório uploads
            file.save(file_path)

            # Processar a planilha e salvar os dados no banco de dados
            save_excel_to_db(file_path)

            # Após salvar no banco, pode remover o arquivo, se não for mais necessário
            os.remove(file_path)

            flash("Planilha carregada e dados inseridos no banco de dados com sucesso!", "success")
            return redirect('/')

        except Exception as e:
            print(f"Erro ao processar a planilha: {e}")
            flash("Ocorreu um erro ao enviar a planilha. Tente novamente mais tarde.", "danger")
            return redirect('/')

    flash("Arquivo inválido. Por favor, envie um arquivo .xlsx.", "danger")
    return redirect('/')

# Rota para baixar a planilha de exemplo
@app.route('/download_example')
def download_example():
    # Criar um arquivo de exemplo
    wb = Workbook()
    ws = wb.active
    ws.title = "Transações Exemplo"
    
    # Cabeçalho da planilha
    ws.append(["ID", "Data", "Quantia", "Descrição", "Valor", "Método de Pagamento", "Tipo"])
    
    # Dados de exemplo
    ws.append([1, "2023-01-01", 100, "Exemplo de Descrição", 50.5, "Pix", "Gasto"])
    
    # Salvar o arquivo em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name="exemplo_transacoes.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/download_relatorio')
def download_relatorio():
    # Conectando ao banco de dados
    with sqlite3.connect('tmp/finance.db') as conn:
        cursor = conn.cursor()

        # Consulta para obter os meses e anos e os totais por método de pagamento
        query = '''
            SELECT strftime('%Y-%m', date) AS month_year,
                   SUM(CASE WHEN payment_method = 'Pix' THEN value ELSE 0 END) AS pix_total,
                   SUM(CASE WHEN payment_method = 'Dinheiro' THEN value ELSE 0 END) AS dinheiro_total,
                   SUM(CASE WHEN payment_method = 'Credito_nu' THEN value ELSE 0 END) AS cartao_nu_total,
                   SUM(CASE WHEN payment_method = 'Credito_c6' THEN value ELSE 0 END) AS cartao_c6_total
            FROM transactions
            WHERE type = 'Gasto'
            GROUP BY month_year
            ORDER BY month_year
        '''
        cursor.execute(query)
        rows = cursor.fetchall()

    # Criando a planilha
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Relatório Financeiro"

    # Cabeçalhos
    headers = ["Data (Mês/Ano)", "Pix", "Dinheiro", "Cartão Nu", "Cartão C6"]
    sheet.append(headers)

    # Adicionando os dados ao Excel
    for row in rows:
        sheet.append(row)

    # Ajustando a largura das colunas
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        column_letter = column[0].column_letter
        sheet.column_dimensions[column_letter].width = max_length + 2

    # Salvando a planilha em um buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Enviando o arquivo para download
    return send_file(
        buffer,
        as_attachment=True,
        download_name="relatorio_financeiro.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/delete/<int:id>', methods=['POST'])
def delete_transaction(id):
    with sqlite3.connect('tmp/finance.db') as conn:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM transactions WHERE id = ?', (id,))
        conn.commit()
    return jsonify({'success': True})

@app.route('/get_transaction/<int:id>', methods=['GET'])
def get_transaction(id):
    with sqlite3.connect('tmp/finance.db') as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM transactions WHERE id = ?', (id,))
        transaction = cursor.fetchone()
    return jsonify(transaction)

@app.route('/edit/<int:id>', methods=['POST'])
def edit_transaction(id):
    data = request.json
    with sqlite3.connect('tmp/finance.db') as conn: 
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE transactions
            SET date = ?, quantia =?, description = ?, value = ?, payment_method = ?, type = ?
            WHERE id = ?
        ''', (data['date'], data['quantia'], data['description'], data['value'], data['payment_method'], data['type'], id))
        conn.commit()
    return jsonify({'success': True})

@app.route('/admin', methods=['GET'])
@login_required
def admin():
    with sqlite3.connect('tmp/finance.db') as conn:
        cursor = conn.cursor()

        # Obter todos os usuários
        cursor.execute('SELECT id, email FROM users')
        users = cursor.fetchall()

    return render_template('admin.html', users=users)

@app.route('/delete_account/<int:id>', methods=['POST'])
@login_required
def delete_account(id):
    # Apagar transações associadas ao usuário
    with sqlite3.connect('tmp/finance.db') as conn:
        cursor = conn.cursor()
        
        # Apagar todas as transações do usuário
        cursor.execute('DELETE FROM transactions WHERE user_id = ?', (id,))
        
        # Apagar o usuário
        cursor.execute('DELETE FROM users WHERE id = ?', (id,))
        conn.commit()

    flash('Conta e dados apagados com sucesso.', 'success')
    return redirect(url_for('admin'))

@app.route('/delete_dados', methods=['GET'])
def delete_dados():
    # Apagar transações associadas ao usuário
    with sqlite3.connect('tmp/finance.db') as conn:
        cursor = conn.cursor()
        
        # Apagar todas as transações do usuário
        cursor.execute('DELETE FROM transactions')  # Corrigido: comando DELETE
        conn.commit()

    return render_template('clear_transactions.html', success=True)

if __name__ == '__main__':
    app.run(debug=True)
